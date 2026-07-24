from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")


def export_screener_report(presets: Dict[str, pd.DataFrame], output_path: Optional[Path | str] = None) -> Path:
    output_path = Path(output_path or "output/screener_output.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    if workbook.sheetnames:
        workbook.remove(workbook.active)

    if not presets:
        sheet = workbook.create_sheet(title="Summary")
        sheet.append(["No data"])
        workbook.save(output_path)
        return output_path

    for preset_name, frame in presets.items():
        sheet_name = preset_name.replace("_", " ").title()[:31]
        sheet = workbook.create_sheet(title=sheet_name)
        if frame is None or frame.empty:
            sheet.append(["No data"])
            continue

        for r_idx, row in enumerate(dataframe_to_rows(frame, index=False, header=True), 1):
            sheet.append(row)

    workbook.save(output_path)
    return output_path


def export_peer_comparison(peer_data: pd.DataFrame, output_path: Optional[Path | str] = None) -> Path:
    output_path = Path(output_path or "output/peer_comparison.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    if workbook.sheetnames:
        workbook.remove(workbook.active)

    if peer_data.empty:
        sheet = workbook.create_sheet(title="Summary")
        sheet.append(["No data"])
        workbook.save(output_path)
        return output_path

    # Pivot data to have one row per company-year and metrics as columns
    pivoted = peer_data.pivot_table(
        index=["company_id", "year", "peer_group"],
        columns="metric",
        values=["value", "rank"]
    ).reset_index()

    # Flatten MultiIndex columns
    pivoted.columns = [f"{c[0]}_{c[1]}" if c[1] else c[0] for c in pivoted.columns]

    for peer_group, group_df in pivoted.groupby("peer_group"):
        sheet_name = str(peer_group).replace("/", "_")[:31]
        sheet = workbook.create_sheet(title=sheet_name)

        # Add median row
        group_df = group_df.copy()
        median_row = group_df.median(numeric_only=True)
        group_df.loc['median'] = median_row

        for r_idx, row_data in enumerate(dataframe_to_rows(group_df, index=False, header=True), 1):
            sheet.append(row_data)

        # Apply formatting
        header = [c.value for c in sheet[1]]
        rank_cols = {c: i for i, c in enumerate(header, 1) if c and c.startswith("rank_")}

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for col_name, col_idx in rank_cols.items():
                cell = row[col_idx - 1]
                if cell.value is None:
                    continue
                
                rank = float(cell.value)
                if rank >= 0.75:
                    cell.fill = GREEN_FILL
                elif rank <= 0.25:
                    cell.fill = RED_FILL
                else:
                    cell.fill = AMBER_FILL

    workbook.save(output_path)
    return output_path
